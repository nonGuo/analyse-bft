import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from .models import ParseResult


def export_to_excel(parse_result: ParseResult, output_path: str):
    rows = []
    for lineage in parse_result.lineage_results:
        for mapping in lineage.mappings:
            rows.append({
                "分组": mapping.group_id,
                "目标Schema": mapping.target_schema,
                "目标表": mapping.target_table,
                "目标字段": mapping.target_column,
                "数据类型": mapping.data_type,
                "来源Schema": ', '.join(mapping.source_schemas),
                "源表": ', '.join(mapping.source_tables),
                "表别名": ', '.join(mapping.source_aliases),
                "源字段": ', '.join(mapping.source_columns),
                "转换规则": mapping.transformation_rule,
                "加工场景": mapping.processing_scenario
            })

    if not rows:
        df = pd.DataFrame(columns=["分组", "目标Schema", "目标表", "目标字段", "数据类型", "来源Schema", "源表", "表别名", "源字段", "转换规则", "加工场景"])
    else:
        df = pd.DataFrame(rows)

    table_lineage_rows = []
    for lineage in parse_result.lineage_results:
        for table_lineage in lineage.table_lineages:
            table_lineage_rows.append({
                "分组": table_lineage.group_id,
                "目标Schema": table_lineage.target_schema,
                "目标表": table_lineage.target_table,
                "来源Schema": table_lineage.source_schema,
                "源表": table_lineage.source_table,
                "表别名": table_lineage.source_alias,
                "关联类型": table_lineage.join_type,
                "关联条件": table_lineage.join_condition,
                "过滤条件": table_lineage.filter_condition,
                "CTE解析": "是" if table_lineage.is_cte_resolved else "否"
            })

    if not table_lineage_rows:
        df_table = pd.DataFrame(columns=["分组", "目标Schema", "目标表", "来源Schema", "源表", "表别名", "关联类型", "关联条件", "过滤条件", "CTE解析"])
    else:
        df_table = pd.DataFrame(table_lineage_rows)

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='字段血缘')
        df_table.to_excel(writer, index=False, sheet_name='表级血缘')

    _format_excel(output_path, '字段血缘')
    _format_excel(output_path, '表级血缘', is_table_lineage=True)


def _format_excel(filepath: str, sheet_name: str = '字段血缘', is_table_lineage: bool = False):
    wb = load_workbook(filepath)
    ws = wb[sheet_name]

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = thin_border

    if is_table_lineage:
        column_widths = {'A': 8, 'B': 20, 'C': 25, 'D': 20, 'E': 25, 'F': 15, 'G': 15, 'H': 50, 'I': 50, 'J': 12}
    else:
        column_widths = {'A': 8, 'B': 20, 'C': 25, 'D': 20, 'E': 15, 'F': 20, 'G': 25, 'H': 15, 'I': 30, 'J': 50, 'K': 15}
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    ws.auto_filter.ref = ws.dimensions

    wb.save(filepath)
